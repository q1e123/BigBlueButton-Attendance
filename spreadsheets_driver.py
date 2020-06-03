# Copyright (C) 2020 Robert-Nicolae Șolcă
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/

from openpyxl import load_workbook

class SpreadSheetDriver():
    def __init__(self,spread_sheet_type, file_name,sheet_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)
        self.ws = self.wb.get_sheet_by_name(sheet_name)

    def save(self):
        self.wb.save(self.file_name)
    def read(self,row,col):
        row = str(row)
        return self.ws[col+row].value

    def write(self,row,col,value):
        row = str(row)
        self.ws[col+row] = value
    def get_row(self,value,col):
        for i in range(1,self.ws.max_row+1):
            if self.read(i,col) == value:
                return i

